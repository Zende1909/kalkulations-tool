"""Excel-Exporte mit openpyxl."""

from __future__ import annotations

import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from app.services.export_models import (
    BaugruppeExportData,
    DashboardExportData,
    ExportTable,
    SpritzgussExportData,
)

EUR_FORMAT = '#,##0.00 "€"'
HEADER_FILL = PatternFill("solid", fgColor="E2E8F0")
BOLD = Font(bold=True)


def _cell_value(value):
    if isinstance(value, datetime) and value.tzinfo is not None:
        return value.replace(tzinfo=None)
    return value


def _autosize(ws) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 50)


def _write_table(ws, start_row: int, export_table: ExportTable, *, table_name: str) -> int:
    row = start_row
    ws.cell(row=row, column=1, value=export_table.title).font = BOLD
    row += 1
    if not export_table.rows:
        ws.cell(row=row, column=1, value="Keine Daten")
        return row + 2
    for c, header in enumerate(export_table.headers, start=1):
        cell = ws.cell(row=row, column=c, value=header)
        cell.font = BOLD
        cell.fill = HEADER_FILL
    row += 1
    first_data = row
    for data_row in export_table.rows:
        for c, val in enumerate(data_row, start=1):
            ws.cell(row=row, column=c, value=_cell_value(val))
        row += 1
    last_row = row - 1
    ref = f"A{first_data - 1}:{get_column_letter(len(export_table.headers))}{last_row}"
    tab = Table(displayName=table_name[:255], ref=ref)
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tab)
    return row + 1


def _write_kv(ws, start_row: int, title: str, rows: list[tuple[str, str]]) -> int:
    row = start_row
    ws.cell(row=row, column=1, value=title).font = BOLD
    row += 1
    for label, value in rows:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)
        row += 1
    return row + 1


def render_spritzguss_excel(data: SpritzgussExportData) -> bytes:
    wb = Workbook()
    ws_over = wb.active
    ws_over.title = "Übersicht"
    ws_over["A1"] = data.company_name
    ws_over["A1"].font = BOLD
    ws_over["A2"] = "Einzelteil-Kalkulation"
    ws_over["A2"].font = BOLD
    meta = [
        ("Kalkulations-ID", data.calculation_id),
        ("Teilenummer", data.teilenummer),
        ("Teilebezeichnung", data.teilebezeichnung),
        ("Kunde", data.kunde),
        ("Projekt", data.projekt),
        ("Endpreis je Stück", data.endpreis),
        ("Erstellt", data.created_at),
        ("Geändert", data.updated_at),
    ]
    r = 4
    for label, val in meta:
        ws_over.cell(row=r, column=1, value=label).font = BOLD
        cell = ws_over.cell(row=r, column=2, value=_cell_value(val))
        if label == "Endpreis je Stück" and isinstance(val, (int, float)):
            cell.number_format = EUR_FORMAT
            cell.font = BOLD
        r += 1
    if data.werkzeug_hinweis:
        ws_over.cell(row=r, column=1, value=data.werkzeug_hinweis).font = Font(color="FF0000", bold=True)
    _autosize(ws_over)

    ws_in = wb.create_sheet("Eingaben")
    _write_kv(ws_in, 1, "Eingabedaten", [(r.label, r.value) for r in data.inputs])
    _autosize(ws_in)

    ws_k = wb.create_sheet("Kostenaufstellung")
    ws_k.cell(row=1, column=1, value="Position").font = BOLD
    ws_k.cell(row=1, column=2, value="Betrag (€)").font = BOLD
    for i, row in enumerate(data.kosten, start=2):
        ws_k.cell(row=i, column=1, value=row.label)
        c = ws_k.cell(row=i, column=2, value=row.amount)
        c.number_format = EUR_FORMAT
        if row.highlight:
            c.font = BOLD
    _autosize(ws_k)

    ws_v = wb.create_sheet("Veredelung")
    if data.veredelung_steps:
        ws_v.cell(row=1, column=1, value="Schritt").font = BOLD
        ws_v.cell(row=1, column=2, value="Kosten").font = BOLD
        for i, step in enumerate(data.veredelung_steps, start=2):
            ws_v.cell(row=i, column=1, value=step.label)
            c = ws_v.cell(row=i, column=2, value=step.amount)
            c.number_format = EUR_FORMAT
    else:
        ws_v["A1"] = "Keine Veredelungsschritte"
    _autosize(ws_v)

    ws_i = wb.create_sheet("Investitionen")
    if data.investitionen:
        headers = ["Bezeichnung", "Typ", "Betrag", "Status", "Hinweis"]
        for c, h in enumerate(headers, 1):
            ws_i.cell(row=1, column=c, value=h).font = BOLD
        for i, inv in enumerate(data.investitionen, start=2):
            ws_i.cell(row=i, column=1, value=inv.bezeichnung)
            ws_i.cell(row=i, column=2, value=inv.typ)
            ws_i.cell(row=i, column=3, value=inv.betrag).number_format = EUR_FORMAT
            ws_i.cell(row=i, column=4, value=inv.status)
            ws_i.cell(row=i, column=5, value=inv.hinweis)
    else:
        ws_i["A1"] = "Keine Investitionen"
    _autosize(ws_i)

    ws_h = wb.create_sheet("Rechenhinweise")
    hints = [
        "Es werden ausschließlich gespeicherte Kalkulationswerte und Preis-Snapshots exportiert.",
        "Der Endpreis je Stück enthält Veredelungskosten als gespeicherten Gesamtwert – keine Doppelzählung.",
        "Einmalinvestitionen für Werkzeuge sind nicht im Stückpreis enthalten.",
    ]
    if data.werkzeug_hinweis:
        hints.append(data.werkzeug_hinweis)
    for i, hint in enumerate(hints, start=1):
        ws_h.cell(row=i, column=1, value=hint)
    _autosize(ws_h)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def render_baugruppe_excel(data: BaugruppeExportData) -> bytes:
    from app.services.baugruppe_export_detail import (
        BaugruppeDetailKalkulation,
        excel_safe_sheet_name,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Deckblatt"
    ws["A1"] = data.company_name
    ws["A1"].font = BOLD
    ws["A2"] = "Baugruppen-Detailkalkulation"
    ws["A2"].font = BOLD
    export_date = data.export_date or datetime.now()
    summary = [
        ("Baugruppen-ID", data.assembly_id),
        ("Name", data.name),
        ("Teilenummer", data.teilenummer),
        ("Kunde", data.kunde),
        ("Programm", data.program or "–"),
        ("Projekt", data.projekt),
        ("Land / Region", data.land or "–"),
        ("Werk / Standort", data.werk or "–"),
        ("Status", data.status or "–"),
        ("Strukturversion", data.structure_version),
        ("Exportdatum", export_date),
        ("Jahresstückzahl", data.jahresstueckzahl),
        ("Herstellkosten", data.herstellkosten),
        ("SG&A / VVGK", data.vvgk),
        ("Selbstkosten", data.selbstkosten),
        ("Profit / Gewinn", data.gewinn),
        ("Skonto", data.skonto),
        ("Nettoverkaufspreis", data.nettoverkaufspreis),
        ("Endpreis je Stück", data.baugruppenpreis_je_stueck),
        ("Jahresumsatz", data.jahresumsatz),
    ]
    r = 4
    money_labels = {
        "Herstellkosten",
        "SG&A / VVGK",
        "Selbstkosten",
        "Profit / Gewinn",
        "Skonto",
        "Nettoverkaufspreis",
        "Endpreis je Stück",
        "Jahresumsatz",
    }
    for label, val in summary:
        ws.cell(row=r, column=1, value=label).font = BOLD
        cell = ws.cell(row=r, column=2, value=_cell_value(val))
        if label in money_labels and isinstance(val, (int, float)):
            cell.number_format = EUR_FORMAT
        if label == "Endpreis je Stück":
            cell.font = BOLD
        r += 1
    # Formeln: Jahresumsatz = Endpreis × Jahresstückzahl (falls beide numerisch gesetzt)
    end_row = next(i for i, (lab, _) in enumerate(summary, start=4) if lab == "Endpreis je Stück")
    jsz_row = next(i for i, (lab, _) in enumerate(summary, start=4) if lab == "Jahresstückzahl")
    umsatz_row = next(i for i, (lab, _) in enumerate(summary, start=4) if lab == "Jahresumsatz")
    if isinstance(data.baugruppenpreis_je_stueck, (int, float)) and data.jahresstueckzahl:
        ws.cell(row=umsatz_row, column=2, value=f"=B{end_row}*B{jsz_row}")
        ws.cell(row=umsatz_row, column=2).number_format = EUR_FORMAT
    _autosize(ws)

    detail = data.detail if isinstance(data.detail, BaugruppeDetailKalkulation) else None

    # Annahmen
    ws_a = wb.create_sheet("Annahmen")
    ws_a["A1"] = "Zentrale Annahmen (aktive Stammdaten)"
    ws_a["A1"].font = BOLD
    headers_a = ["Bezeichnung", "Satz %", "Kostenbasis", "Hinweis"]
    for c, h in enumerate(headers_a, 1):
        cell = ws_a.cell(row=3, column=c, value=h)
        cell.font = BOLD
        cell.fill = HEADER_FILL
    if detail and detail.assumptions:
        for i, a in enumerate(detail.assumptions, start=4):
            ws_a.cell(row=i, column=1, value=a.bezeichnung)
            ws_a.cell(row=i, column=2, value=a.satz_prozent)
            ws_a.cell(row=i, column=3, value=a.kostenbasis)
            ws_a.cell(row=i, column=4, value=a.hinweis)
    else:
        ws_a["A4"] = "Keine Annahmen verfügbar"
    _autosize(ws_a)

    if data.bom is not None:
        ws_bom = wb.create_sheet("BOM")
        _write_table(ws_bom, 1, data.bom, table_name="BOM")
        _autosize(ws_bom)

    ws_e = wb.create_sheet("Einzelteile")
    _write_table(ws_e, 1, data.einzelteile, table_name="Einzelteile")
    _autosize(ws_e)

    used_sheets: set[str] = set(wb.sheetnames)

    def _kv_money(ws, row: int, label: str, value) -> int:
        ws.cell(row=row, column=1, value=label).font = BOLD
        cell = ws.cell(row=row, column=2, value=value if value is not None else "–")
        if isinstance(value, (int, float)):
            cell.number_format = EUR_FORMAT
        return row + 1

    if detail:
        for part in detail.parts:
            base = f"PART_{part.sequence:03d}_{part.sheet_slug or part.bezeichnung}"
            sheet_name = excel_safe_sheet_name(base, used_sheets)
            ws_p = wb.create_sheet(sheet_name)
            ws_p["A1"] = f"Detailkalkulation: {part.bezeichnung}"
            ws_p["A1"].font = BOLD
            r = 3
            meta = [
                ("Bezeichnung", part.bezeichnung),
                ("Teilenummer", part.teilenummer),
                ("Menge", part.menge),
                ("Kostenbasis", part.price_basis),
                ("Material", part.material_name or "–"),
                ("Materialpreis €/kg", part.materialpreis_pro_kg),
                ("Netto-Teilegewicht g", part.teilegewicht_netto_g),
                ("Schussgewicht g", part.schussgewicht_g),
                ("Materialkosten direkt", part.materialkosten),
                ("Material-Ausschuss %", part.material_ausschussquote_pct),
                ("Material inkl. Ausschuss", part.materialkosten_inkl_ausschuss),
                ("Material-Nominierung", part.material_nominierung or "fehlend"),
                ("MGK %", part.mgk_pct),
                ("Material-MGK", part.material_mgk),
                ("Maschinenkosten", part.maschinenkosten),
                ("Fertigungslohn", part.fertigungslohn),
                ("Spritzguss-Ausgangskosten", part.spritzguss_ausgang),
                ("Veredelung direkt vor Ausschuss", part.veredelung_direkt_vor),
                ("FGK-Basis", part.fgk_basis),
                ("FGK %", part.fgk_pct),
                ("FGK-Betrag", part.fgk_betrag),
                ("Herstellkosten", part.herstellkosten),
                ("Zwischensumme", part.zwischensumme),
            ]
            for label, val in meta:
                ws_p.cell(row=r, column=1, value=label).font = BOLD
                cell = ws_p.cell(row=r, column=2, value=val if val is not None else "–")
                if isinstance(val, (int, float)) and label not in {
                    "Menge",
                    "Material-Ausschuss %",
                    "MGK %",
                    "FGK %",
                    "Netto-Teilegewicht g",
                    "Schussgewicht g",
                }:
                    cell.number_format = EUR_FORMAT
                r += 1
            if part.hinweise:
                r += 1
                ws_p.cell(row=r, column=1, value="Hinweise").font = BOLD
                r += 1
                for h in part.hinweise:
                    ws_p.cell(row=r, column=1, value=h)
                    r += 1
            _autosize(ws_p)

            # Eigene Prozesskette je PART (auch wenn leer: Spritzguss-Ausgang)
            pk_name = excel_safe_sheet_name(
                f"PART_{part.sequence:03d}_Prozesskette", used_sheets
            )
            ws_pk = wb.create_sheet(pk_name)
            ws_pk["A1"] = f"Prozesskette: {part.bezeichnung}"
            ws_pk["A1"].font = BOLD
            headers = [
                "Reihenfolge",
                "Prozess",
                "Art",
                "Lohn",
                "Maschine",
                "Verbrauch",
                "Direkt vor Ausschuss",
                "Ausschuss %",
                "Ausbeute %",
                "Vorprodukt",
                "Ausschusszuschlag",
                "Nach Ausbeute",
            ]
            for c, h in enumerate(headers, 1):
                cell = ws_pk.cell(row=3, column=c, value=h)
                cell.font = BOLD
                cell.fill = HEADER_FILL
            pk_row = 4
            ws_pk.cell(row=pk_row, column=1, value=0)
            ws_pk.cell(row=pk_row, column=2, value="Spritzguss Ausgang")
            ws_pk.cell(row=pk_row, column=3, value="Spritzguss")
            ws_pk.cell(row=pk_row, column=7, value=part.spritzguss_ausgang).number_format = EUR_FORMAT
            ws_pk.cell(row=pk_row, column=8, value=part.material_ausschussquote_pct)
            ws_pk.cell(row=pk_row, column=10, value=part.materialkosten).number_format = EUR_FORMAT
            if part.materialkosten_inkl_ausschuss is not None and part.materialkosten is not None:
                ws_pk.cell(
                    row=pk_row,
                    column=11,
                    value=part.materialkosten_inkl_ausschuss - part.materialkosten,
                ).number_format = EUR_FORMAT
            ws_pk.cell(
                row=pk_row, column=12, value=part.materialkosten_inkl_ausschuss
            ).number_format = EUR_FORMAT
            pk_row += 1
            for step in part.process_steps:
                ws_pk.cell(row=pk_row, column=1, value=step.reihenfolge)
                ws_pk.cell(row=pk_row, column=2, value=step.bezeichnung)
                ws_pk.cell(row=pk_row, column=3, value=step.veredelungsart)
                ws_pk.cell(row=pk_row, column=4, value=step.lohnkosten).number_format = EUR_FORMAT
                ws_pk.cell(row=pk_row, column=5, value=step.maschinenkosten).number_format = EUR_FORMAT
                ws_pk.cell(row=pk_row, column=6, value=step.verbrauchskosten).number_format = EUR_FORMAT
                ws_pk.cell(row=pk_row, column=7, value=step.kosten_vor_ausschuss).number_format = EUR_FORMAT
                ws_pk.cell(row=pk_row, column=8, value=step.ausschussquote_pct)
                ws_pk.cell(row=pk_row, column=9, value=step.ausbeute_pct)
                ws_pk.cell(row=pk_row, column=10, value=step.vorprodukt_eingang).number_format = EUR_FORMAT
                ws_pk.cell(row=pk_row, column=11, value=step.ausschuss_zuschlag).number_format = EUR_FORMAT
                ws_pk.cell(row=pk_row, column=12, value=step.kosten_nach_ausbeute).number_format = EUR_FORMAT
                pk_row += 1
            if part.fgk_betrag is not None:
                pk_row += 1
                ws_pk.cell(row=pk_row, column=1, value="FGK").font = BOLD
                ws_pk.cell(row=pk_row, column=2, value=f"Satz {part.fgk_pct} %")
                ws_pk.cell(row=pk_row, column=10, value=part.fgk_basis).number_format = EUR_FORMAT
                ws_pk.cell(row=pk_row, column=12, value=part.fgk_betrag).number_format = EUR_FORMAT
                pk_row += 1
                ws_pk.cell(row=pk_row, column=1, value="Herstellkosten fertig").font = BOLD
                ws_pk.cell(row=pk_row, column=12, value=part.herstellkosten).number_format = EUR_FORMAT
            _autosize(ws_pk)

    ws_k = wb.create_sheet(excel_safe_sheet_name("Kaufteile_Detail", used_sheets) if detail else "Kaufteile")
    _write_table(ws_k, 1, data.kaufteile, table_name="Kaufteile")
    _autosize(ws_k)

    ws_v = wb.create_sheet(excel_safe_sheet_name("ASSY_Detail", used_sheets) if detail else "ASSY_Prozesskette")
    _write_table(ws_v, 1, data.veredelung, table_name="ASSY")
    if detail and detail.processes:
        extra = len(data.veredelung.rows) + 4
        ws_v.cell(row=extra, column=1, value="Vorprodukte je ASSY-Prozess").font = BOLD
        extra += 1
        for proc in detail.processes:
            ws_v.cell(
                row=extra,
                column=1,
                value=f"{proc.bezeichnung}: "
                + (
                    ", ".join(proc.vorprodukt_komponenten)
                    if proc.vorprodukt_komponenten
                    else "–"
                ),
            )
            extra += 1
    _autosize(ws_v)

    ws_i = wb.create_sheet("Investitionen")
    if data.investitionen:
        headers = ["Bezeichnung", "Typ", "Betrag", "Status", "Hinweis"]
        for c, h in enumerate(headers, 1):
            ws_i.cell(row=1, column=c, value=h).font = BOLD
        for i, inv in enumerate(data.investitionen, start=2):
            ws_i.cell(row=i, column=1, value=inv.bezeichnung)
            ws_i.cell(row=i, column=2, value=inv.typ)
            ws_i.cell(row=i, column=3, value=inv.betrag).number_format = EUR_FORMAT
            ws_i.cell(row=i, column=4, value=inv.status)
            ws_i.cell(row=i, column=5, value=inv.hinweis or "Separat, nicht im Stückpreis enthalten")
    else:
        ws_i["A1"] = "Keine Investitionen – separat, nicht im Stückpreis enthalten"
    _autosize(ws_i)

    if data.zuschlagssaetze is not None:
        ws_m = wb.create_sheet("Zuschlagssaetze")
        _write_table(ws_m, 1, data.zuschlagssaetze, table_name="Zuschlaege")
        _autosize(ws_m)

    ws_u = wb.create_sheet(
        excel_safe_sheet_name("Gesamtueberleitung", used_sheets) if detail else "Ueberleitung"
    )
    ws_u["A1"] = "Gesamtüberleitung zum Endpreis"
    ws_u["A1"].font = BOLD
    for c, h in enumerate(["Position", "Betrag", "Berechnungsbasis"], 1):
        cell = ws_u.cell(row=3, column=c, value=h)
        cell.font = BOLD
        cell.fill = HEADER_FILL
    u_row = 4
    if detail and detail.ueberleitung:
        for line in detail.ueberleitung:
            ws_u.cell(row=u_row, column=1, value=line.label)
            if line.amount is not None:
                ws_u.cell(row=u_row, column=2, value=line.amount).number_format = EUR_FORMAT
            else:
                ws_u.cell(row=u_row, column=2, value="–")
            if line.highlight:
                ws_u.cell(row=u_row, column=1).font = BOLD
                ws_u.cell(row=u_row, column=2).font = BOLD
            ws_u.cell(row=u_row, column=3, value=line.basis)
            u_row += 1
    else:
        ws_u["A4"] = "Keine Überleitung"
    _autosize(ws_u)

    ws_z = wb.create_sheet("Zusammenfassung")
    totals = [
        ("Einzelteile gesamt", data.einzelteile_gesamt),
        ("Kaufteile gesamt", data.kaufteile_gesamt),
        ("Veredelung/ASSY gesamt", data.veredelung_gesamt),
        ("Herstellkosten", data.herstellkosten),
        ("SG&A / VVGK", data.vvgk),
        ("Selbstkosten", data.selbstkosten),
        ("Gewinn", data.gewinn),
        ("Skonto", data.skonto),
        ("Nettoverkaufspreis", data.nettoverkaufspreis),
        ("Preis pro Stück", data.baugruppenpreis_je_stueck),
        ("Jahresstückzahl", data.jahresstueckzahl),
        ("Jahresumsatz", data.jahresumsatz),
        ("Gesamtergebnis", data.gesamtergebnis),
    ]
    row_i = 1
    if data.kosten_aufstellung:
        ws_z.cell(row=row_i, column=1, value="Kostenaufstellung").font = BOLD
        row_i += 1
        for item in data.kosten_aufstellung:
            ws_z.cell(row=row_i, column=1, value=item.label)
            c = ws_z.cell(row=row_i, column=2, value=item.amount)
            c.number_format = EUR_FORMAT
            if item.highlight:
                c.font = BOLD
            row_i += 1
        row_i += 1
    for label, val in totals:
        highlight = label in {"Preis pro Stück", "Gesamtergebnis", "Jahresumsatz"}
        ws_z.cell(row=row_i, column=1, value=label).font = BOLD if highlight else None
        c = ws_z.cell(row=row_i, column=2, value=val)
        if label != "Jahresstückzahl":
            c.number_format = EUR_FORMAT
        if highlight:
            c.font = BOLD
        row_i += 1
    if detail and detail.jahresstueckzahl_hinweis:
        row_i += 1
        ws_z.cell(row=row_i, column=1, value=detail.jahresstueckzahl_hinweis)
        row_i += 1
    if detail and detail.warnings:
        row_i += 1
        ws_z.cell(row=row_i, column=1, value="Hinweise / Datenqualität").font = BOLD
        row_i += 1
        for w in detail.warnings:
            ws_z.cell(row=row_i, column=1, value=w)
            row_i += 1
    _autosize(ws_z)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def render_dashboard_excel(data: DashboardExportData) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "KPI-Übersicht"
    ws["A1"] = data.company_name
    ws["A1"].font = BOLD
    ws["A2"] = "Dashboard-Bericht"
    ws["A3"] = f"Erstellt: {data.generated_at.strftime('%d.%m.%Y %H:%M')}"
    filt = []
    if data.filter_project:
        filt.append(f"Projekt: {data.filter_project}")
    if data.filter_customer:
        filt.append(f"Kunde: {data.filter_customer}")
    if data.filter_status:
        filt.append(f"Status: {data.filter_status}")
    if data.filter_date_from or data.filter_date_to:
        filt.append(f"Zeitraum: {data.filter_date_from or '–'} bis {data.filter_date_to or '–'}")
    if data.filter_kalkulationsart:
        filt.append(f"Kalkulationsart: {data.filter_kalkulationsart}")
    ws["A4"] = "Filter: " + (", ".join(filt) if filt else "Keine")
    if data.empty_message:
        ws["A5"] = data.empty_message
    r = 6
    for kpi in data.kpis:
        ws.cell(row=r, column=1, value=kpi.label).font = BOLD
        ws.cell(row=r, column=2, value=kpi.value)
        r += 1
    _autosize(ws)

    sheets = [
        ("Kalkulationen", data.recent_calculations, "Kalkulationen"),
        ("Baugruppen", data.assemblies, "Baugruppen"),
        ("Investitionen", data.investments, "Investitionen"),
        ("Umsatzpotenzial", data.revenue_by_project, "Umsatz"),
        ("Preisvergleich", data.price_comparison, "Preise"),
    ]
    for sheet_name, table, tname in sheets:
        ws_t = wb.create_sheet(sheet_name)
        _write_table(ws_t, 1, table, table_name=tname)
        _autosize(ws_t)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
