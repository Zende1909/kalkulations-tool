"""Excel/PDF-Export für Business-Case-Übersicht."""

from __future__ import annotations

import io
from dataclasses import dataclass, field
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy.orm import Session

from app.config import settings
from app.services.business_case_overview import build_project_business_case

EUR_FORMAT = '#,##0.00 "€"'
PCT_FORMAT = '0.00"%"'
BOLD = Font(bold=True)


@dataclass
class BusinessCaseExportData:
    company_name: str
    customer: str
    program: str
    project: str
    generated_at: datetime
    kpis: dict
    parts: list[dict]
    assemblies: list[dict]
    investments: list[dict]
    investment_financial_summary: dict = field(default_factory=dict)
    position_headers: list[str] = field(default_factory=list)
    position_rows: list[list] = field(default_factory=list)
    investment_headers: list[str] = field(default_factory=list)
    investment_rows: list[list] = field(default_factory=list)


def _revenue_export(value: float | None) -> int | None:
    if value is None:
        return None
    return int(round(value))


def _money(value: float | None) -> float | str | None:
    return value


def _pct(value: float | None) -> float | str | None:
    return value


def _position_row(item: dict, typ: str, label_key: str) -> list:
    return [
        typ,
        item.get("material_number") or item.get("teilenummer"),
        item.get(label_key),
        _money(item.get("cost_per_piece")),
        _money(item.get("bottom_price_per_piece")),
        _money(item.get("actual_price_per_piece")),
        _money(item.get("guide_price_per_piece")),
        item.get("project_volume"),
        _revenue_export(item.get("bottom_price_revenue")),
        _revenue_export(item.get("actual_revenue")),
        _money(item.get("cost_total")),
        _money(item.get("margin_bottom_price_total")),
        _pct(item.get("margin_bottom_price_total_pct")),
        _money(item.get("margin_actual_total")),
        _pct(item.get("margin_actual_total_pct")),
        _pct(item.get("margin_bottom_price_pct")),
        _pct(item.get("margin_actual_price_pct")),
    ]


def build_business_case_export(
    db: Session,
    *,
    customer_id: int,
    program_id: int,
    linked_project_id: int,
) -> BusinessCaseExportData:
    data = build_project_business_case(
        db,
        customer_id=customer_id,
        program_id=program_id,
        linked_project_id=linked_project_id,
    )
    position_headers = [
        "Typ",
        "Materialnummer",
        "Bezeichnung",
        "Kosten/Stück",
        "Bottom Price/Stück",
        "Tatsächlicher Preis/Stück",
        "Richtpreis (15 %)",
        "Projektstückzahl",
        "Bottom-Price-Umsatz",
        "Tatsächlicher Umsatz",
        "Kosten gesamt",
        "Bottom-Marge gesamt",
        "Bottom-Marge gesamt %",
        "Tatsächliche Marge gesamt",
        "Tatsächliche Marge gesamt %",
        "Bottom-Marge/Stück %",
        "Tatsächliche Marge/Stück %",
    ]
    position_rows: list[list] = []
    for part in data["parts"]:
        position_rows.append(_position_row(part, "Einzelteil", "bezeichnung"))
    for asm in data["assemblies"]:
        position_rows.append(_position_row(asm, "Baugruppe", "name"))

    investment_headers = [
        "Zahlungsart",
        "Bezeichnung",
        "Zuordnungstyp",
        "Materialnummer",
        "Kunde",
        "Programm",
        "Projekt",
        "Kosten einmalig",
        "Bottom Price einmalig",
        "Erlös einmalig",
        "Erlös − Kosten",
        "Erlös − Kosten %",
        "Erlös − Bottom Price",
        "Erlös − Bottom Price %",
        "Bottom Price − Kosten",
    ]
    investment_rows = [
        [
            inv.get("payment_type"),
            inv.get("bezeichnung"),
            inv.get("assignment_type_label") or inv.get("assignment_type"),
            inv.get("material_number") or "",
            inv.get("customer_name"),
            inv.get("program_name"),
            inv.get("project_name"),
            inv.get("cost_amount"),
            inv.get("bottom_price"),
            inv.get("revenue_amount"),
            inv.get("margin_revenue_minus_cost"),
            _pct(inv.get("margin_revenue_minus_cost_pct")),
            inv.get("margin_revenue_minus_bottom_price"),
            _pct(inv.get("margin_revenue_minus_bottom_price_pct")),
            inv.get("margin_bottom_price_minus_cost"),
        ]
        for inv in data["investments"]
    ]

    return BusinessCaseExportData(
        company_name=settings.COMPANY_NAME,
        customer=data["customer"],
        program=data.get("program", ""),
        project=data["project"],
        generated_at=datetime.now(timezone.utc),
        kpis=data["kpis"],
        parts=data["parts"],
        assemblies=data["assemblies"],
        investments=data["investments"],
        investment_financial_summary=data.get("investment_financial_summary") or {},
        position_headers=position_headers,
        position_rows=position_rows,
        investment_headers=investment_headers,
        investment_rows=investment_rows,
    )


def render_business_case_excel(data: BusinessCaseExportData) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Business Case"
    ws["A1"] = f"{data.company_name} – Business Case"
    ws["A1"].font = BOLD
    ws["A2"] = f"{data.customer} / {data.program} / {data.project}"
    row = 4
    kpi_labels = [
        ("Operative Kosten", data.kpis.get("operative_cost_total"), EUR_FORMAT),
        ("Gebundenes Projektkapital", data.kpis.get("bound_capital_total"), EUR_FORMAT),
        ("CAPEX gesamt", data.kpis.get("capex_cost_total"), EUR_FORMAT),
        ("Sonstige Investitionskosten", data.kpis.get("non_capex_investment_cost_total"), EUR_FORMAT),
        ("Gesamtinvestitionskosten", data.kpis.get("investition_cost_total"), EUR_FORMAT),
        ("Teileumsatz Bottom Price", _revenue_export(data.kpis.get("parts_bottom_price_revenue_total")), None),
        ("Teileumsatz tatsächlich", _revenue_export(data.kpis.get("parts_actual_revenue_total")), None),
        ("Investitionserlös Bottom Price", _revenue_export(data.kpis.get("investition_bottom_price_total")), None),
        ("Investitionserlös tatsächlich", _revenue_export(data.kpis.get("investition_revenue_total")), None),
        ("Gesamtumsatz Bottom Price", _revenue_export(data.kpis.get("bottom_price_revenue_total")), None),
        ("Gesamtumsatz tatsächlich", _revenue_export(data.kpis.get("actual_revenue_total")), None),
        ("EBIT Bottom Price", data.kpis.get("ebit_bottom_total"), EUR_FORMAT),
        ("EBIT Bottom Price %", data.kpis.get("ebit_bottom_total_pct"), PCT_FORMAT),
        ("EBIT tatsächlich", data.kpis.get("ebit_actual_total"), EUR_FORMAT),
        ("EBIT tatsächlich %", data.kpis.get("ebit_actual_total_pct"), PCT_FORMAT),
        ("ROI Bottom Price inkl. CAPEX %", data.kpis.get("roi_incl_capex_bottom_pct"), PCT_FORMAT),
        ("ROI tatsächlich inkl. CAPEX %", data.kpis.get("roi_incl_capex_actual_pct"), PCT_FORMAT),
        ("Operativer ROI Bottom Price %", data.kpis.get("roi_operating_bottom_pct"), PCT_FORMAT),
        ("Operativer ROI tatsächlich %", data.kpis.get("roi_operating_actual_pct"), PCT_FORMAT),
        ("Projektstückzahl", data.kpis.get("project_volume_total"), None),
        ("Einzelteile", data.kpis.get("anzahl_einzelteile"), None),
        ("Baugruppen", data.kpis.get("anzahl_baugruppen"), None),
    ]
    for label, val, fmt in kpi_labels:
        ws.cell(row=row, column=1, value=label).font = BOLD
        cell = ws.cell(row=row, column=2, value=val)
        if isinstance(val, float) and fmt:
            cell.number_format = fmt
        elif isinstance(val, int) and label.endswith("Umsatz"):
            cell.number_format = '#,##0 "€"'
        row += 1
    row += 1
    ws_pos = wb.create_sheet("Materialpositionen")
    position_pct_cols = {13, 15, 16, 17}
    position_revenue_cols = {9, 10}
    for col, header in enumerate(data.position_headers, 1):
        ws_pos.cell(row=1, column=col, value=header).font = BOLD
    for i, prow in enumerate(data.position_rows, start=2):
        for col, val in enumerate(prow, 1):
            cell = ws_pos.cell(row=i, column=col, value=val)
            if isinstance(val, float):
                cell.number_format = PCT_FORMAT if col in position_pct_cols else EUR_FORMAT
            elif col in position_revenue_cols and isinstance(val, int):
                cell.number_format = '#,##0 "€"'
            elif col == 8 and isinstance(val, (int, float)):
                cell.number_format = "#,##0"
    ws_inv = wb.create_sheet("Investitionen")
    inv_pct_cols = {12, 14}
    for col, header in enumerate(data.investment_headers, 1):
        ws_inv.cell(row=1, column=col, value=header).font = BOLD
    for i, irow in enumerate(data.investment_rows, start=2):
        for col, val in enumerate(irow, 1):
            cell = ws_inv.cell(row=i, column=col, value=val)
            if isinstance(val, float):
                cell.number_format = PCT_FORMAT if col in inv_pct_cols else EUR_FORMAT
    summary_row = len(data.investment_rows) + 3
    ws_inv.cell(row=summary_row, column=1, value="Summen").font = BOLD
    fin = data.investment_financial_summary or {}
    summary_labels = [
        ("CAPEX / Werksinvestitionen", fin.get("capex", {})),
        ("Entwicklungsinvestitionen", fin.get("entwicklung", {})),
        ("Amortisation / Einmalzahlung", fin.get("legacy", {})),
        ("Gesamt Investitionskosten", fin.get("totals", {})),
    ]
    for offset, (label, block) in enumerate(summary_labels, start=1):
        row_idx = summary_row + offset
        ws_inv.cell(row=row_idx, column=1, value=label)
        ws_inv.cell(row=row_idx, column=2, value=(block or {}).get("count"))
        cost_cell = ws_inv.cell(row=row_idx, column=8, value=(block or {}).get("cost_amount_total"))
        cost_cell.number_format = EUR_FORMAT
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
