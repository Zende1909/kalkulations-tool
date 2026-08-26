"""Idempotenter Seed: Saudi-Arabien / KAEC aus Mappe1.xlsx.

Nur CLI – nie App-Startup, nie Alembic-DML.

Aufruf (aus backend/):

    python -m app.scripts.seed_kaec_costing_from_mappe1

Quelle
------
- Datei: backend/data/reference/Mappe1.xlsx
- Blatt: Costing_Base_Data
- Wechselkurs: 1 USD = 0.92 EUR

Verhalten
---------
- Legt fehlende Länder/Werke/Maschinen/Löhne/Zuschläge an.
- Vorhandene Datensätze (Unique Keys) werden nicht überschrieben → Report.
- Excel „Overhead raw material 2 %“ wird als inaktiver Hinweis importiert;
  die verbindliche Material-MGK bleibt 3 %/5 % (Nominierung).
"""

from __future__ import annotations

import sys
from datetime import date
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.config import Settings, settings
from app.database import SessionLocal
from app.models.land import Land
from app.models.lohnkosten import Lohnkosten
from app.models.maschine import Maschine
from app.models.werk import Werk
from app.models.werk_zuschlag import WerkZuschlag
from app.models.user import utcnow
from app.services.machine_hourly_rate import (
    MachineRateInput,
    apply_rate_to_maschine,
    berechne_maschinenstundensatz,
)

REFERENCE_XLSX = (
    Path(__file__).resolve().parents[2] / "data" / "reference" / "Mappe1.xlsx"
)
SHEET = "Costing_Base_Data"
FX_USD_TO_EUR = 0.92
LAND_CODE = "SA"
LAND_NAME = "Saudi-Arabien"
WERK_CODE = "KAEC"
WERK_NAME = "KAEC"


def is_local_development_database(database_url: str) -> bool:
    return Settings.is_local_development_database_url(database_url)


def _cell(ws, row: int, col: int):
    return ws.cell(row, col).value


def _num(value, default: float = 0.0) -> float:
    if value is None:
        return default
    if isinstance(value, str) and value.startswith("="):
        # Formel – für Seed data_only=True nutzen; Fallback 0
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _resolve_formula_number(ws_data, ws_form, row: int, col: int, default: float = 0.0) -> float:
    """Bevorzugt gecachte Excel-Werte; sonst einfache Literal-Formeln wie =x*y."""
    cached = ws_data.cell(row, col).value
    if cached is not None and not isinstance(cached, str):
        return float(cached)
    raw = ws_form.cell(row, col).value
    if isinstance(raw, (int, float)):
        return float(raw)
    if isinstance(raw, str) and raw.startswith("=") and "*" in raw and raw.count("*") == 1:
        # z. B. =1900000*1.15
        left, right = raw[1:].split("*", 1)
        try:
            return float(left.strip()) * float(right.strip())
        except ValueError:
            return default
    return default


def seed_kaec_from_mappe1(db: Session, xlsx_path: Path | None = None) -> list[str]:
    path = xlsx_path or REFERENCE_XLSX
    if not path.is_file():
        raise FileNotFoundError(f"Referenzdatei fehlt: {path}")

    actions: list[str] = []
    wb_form = load_workbook(path, data_only=False)
    wb_data = load_workbook(path, data_only=True)
    ws_f = wb_form[SHEET]
    ws_d = wb_data[SHEET]

    # --- Land / Werk ---
    land = db.scalars(select(Land).where(Land.code == LAND_CODE)).first()
    if land is None:
        land = Land(code=LAND_CODE, name=LAND_NAME, aktiv=True)
        db.add(land)
        db.flush()
        actions.append(f"insert:land:{LAND_CODE}")
    else:
        actions.append(f"skip:land:{LAND_CODE}")

    werk = db.scalars(select(Werk).where(Werk.code == WERK_CODE)).first()
    if werk is None:
        werk = Werk(
            land_id=land.id,
            code=WERK_CODE,
            name=WERK_NAME,
            currency="USD",
            fx_to_eur=FX_USD_TO_EUR,
            aktiv=True,
        )
        db.add(werk)
        db.flush()
        actions.append(f"insert:werk:{WERK_CODE}")
    else:
        actions.append(f"skip:werk:{WERK_CODE}")

    # Globals Zeile 3
    days = _num(_cell(ws_d, 3, 7), 254)
    shifts = _num(_cell(ws_d, 3, 8), 2)
    hours = _num(_cell(ws_d, 3, 9), 8)
    oee = _num(_cell(ws_d, 3, 10), 0.9)
    space_satz = _num(_cell(ws_d, 3, 14), 30)
    depr_years = _num(_cell(ws_d, 3, 17), 10)
    interest = _num(_cell(ws_d, 3, 21), 0.08)
    insurance = _num(_cell(ws_d, 3, 24), 0.0045)
    maint = _num(_cell(ws_d, 3, 27), 0.02)
    power_price = _num(_cell(ws_d, 3, 31), 0.06)
    air_price = _num(_cell(ws_d, 3, 34), 0.06)
    water_price = _num(_cell(ws_d, 3, 37), 0.03)

    # Maschinenzeilen 6–37
    for row in range(6, 38):
        typ = _cell(ws_f, row, 4)
        var = _cell(ws_f, row, 5)
        if typ is None or str(typ).strip() == "":
            continue
        typ_s = str(typ).strip()
        var_s = str(var).strip() if var is not None else ""
        nr = f"KAEC-{typ_s.replace(' ', '')}-{var_s}".replace("+", "")[:50]
        existing = db.scalars(select(Maschine).where(Maschine.maschinen_nr == nr)).first()
        if existing is not None:
            actions.append(f"skip:maschine:{nr}")
            continue

        investment = _resolve_formula_number(ws_d, ws_f, row, 12)
        flaeche = _num(_cell(ws_d, row, 13) or _cell(ws_f, row, 13))
        setup_min = _num(_cell(ws_d, row, 41) or _cell(ws_f, row, 41), 45)
        setup_workers = _num(_cell(ws_d, row, 43) or _cell(ws_f, row, 43), 1)
        strom_v = _num(_cell(ws_d, row, 30) or _cell(ws_f, row, 30))
        druck_v = _num(_cell(ws_d, row, 33) or _cell(ws_f, row, 33))
        kuehl_v = _num(_cell(ws_d, row, 36) or _cell(ws_f, row, 36))

        # Schließkraft: numerische Variante oder 0
        try:
            schliess = float(str(var_s).split()[0])
        except (ValueError, IndexError):
            schliess = 0.0

        rate_in = MachineRateInput(
            arbeitstage_pro_jahr=days,
            schichten_pro_tag=shifts,
            stunden_pro_schicht=hours,
            oee=oee,
            investment=investment,
            flaeche_sqm=flaeche,
            space_cost_satz_pro_sqm_jahr=space_satz,
            abschreibungsdauer_jahre=depr_years,
            zinssatz=interest,
            versicherungssatz=insurance,
            instandhaltungssatz=maint,
            stromverbrauch_kwh_h=strom_v,
            strompreis=power_price,
            druckluftverbrauch_m3_h=druck_v,
            druckluftpreis=air_price,
            kuehlwasserverbrauch_m3_h=kuehl_v,
            kuehlwasserpreis=water_price,
            fx_to_eur=FX_USD_TO_EUR,
            source_currency="USD",
        )
        rate = berechne_maschinenstundensatz(rate_in)
        m = Maschine(
            bezeichnung=f"{typ_s} {var_s}".strip(),
            maschinen_nr=nr,
            stundensatz=rate.stundensatz_eur,
            schliesskraft_t=schliess,
            aktiv=True,
            werk_id=werk.id,
            maschinentyp=typ_s,
            variante=var_s,
            source_currency="USD",
            arbeitstage_pro_jahr=days,
            schichten_pro_tag=shifts,
            stunden_pro_schicht=hours,
            oee=oee,
            investment=investment,
            flaeche_sqm=flaeche,
            space_cost_satz_pro_sqm_jahr=space_satz,
            abschreibungsdauer_jahre=depr_years,
            zinssatz=interest,
            versicherungssatz=insurance,
            instandhaltungssatz=maint,
            stromverbrauch_kwh_h=strom_v,
            strompreis=power_price,
            druckluftverbrauch_m3_h=druck_v,
            druckluftpreis=air_price,
            kuehlwasserverbrauch_m3_h=kuehl_v,
            kuehlwasserpreis=water_price,
            setup_zeit_min=setup_min,
            setup_mitarbeiter=setup_workers,
            rate_updated_at=utcnow(),
        )
        apply_rate_to_maschine(m, rate)
        db.add(m)
        actions.append(f"insert:maschine:{nr}")

    # Löhne
    for rolle, usd, label in (
        ("produktion", 12.0, "KAEC Produktionsmitarbeiter"),
        ("setup", 25.0, "KAEC Setup-Mitarbeiter"),
    ):
        existing = db.scalars(
            select(Lohnkosten).where(
                Lohnkosten.werk_id == werk.id,
                Lohnkosten.rolle == rolle,
                Lohnkosten.aktiv.is_(True),
            )
        ).first()
        if existing is not None:
            actions.append(f"skip:lohn:{rolle}")
            continue
        db.add(
            Lohnkosten(
                bezeichnung=label,
                kosten_pro_stunde=round(usd * FX_USD_TO_EUR, 6),
                kostenstelle=f"KAEC-{rolle}",
                gueltig_ab=date(2026, 1, 1),
                aktiv=True,
                werk_id=werk.id,
                rolle=rolle,
                source_currency="USD",
                source_rate=usd,
            )
        )
        actions.append(f"insert:lohn:{rolle}")

    # Werk-Zuschläge
    zuschlaege = [
        ("fgk", "Overhead production / FGK", 22.0, "fertigung", True),
        ("vvgk", "Overhead SG&A", 10.0, "herstellkosten", True),
        ("gewinn", "Profit charge", 15.0, "selbstkosten", True),
        ("handling_oem_kaufteil", "Handling Charge OEM parts", 6.0, "einkaufspreis", True),
        (
            "overhead_raw_material_excel",
            "Excel Overhead raw material (NICHT angewendet – Konflikt 3%/5% MGK)",
            2.0,
            "material",
            False,
        ),
    ]
    for typ, bez, satz, basis, aktiv in zuschlaege:
        existing = db.scalars(
            select(WerkZuschlag).where(
                WerkZuschlag.werk_id == werk.id, WerkZuschlag.typ == typ
            )
        ).first()
        if existing is not None:
            actions.append(f"skip:zuschlag:{typ}")
            continue
        db.add(
            WerkZuschlag(
                werk_id=werk.id,
                typ=typ,
                bezeichnung=bez,
                satz_prozent=satz,
                kostenbasis=basis,
                aktiv=aktiv,
            )
        )
        actions.append(f"insert:zuschlag:{typ}")
        if typ == "overhead_raw_material_excel":
            actions.append(
                "conflict:material_mgk:excel_2pct_skipped_keep_central_3_5"
            )

    db.commit()
    return actions


def main() -> int:
    if not is_local_development_database(settings.DATABASE_URL):
        print("Abbruch: Seed nur für lokale Entwicklungsdatenbanken.", file=sys.stderr)
        return 2
    db = SessionLocal()
    try:
        actions = seed_kaec_from_mappe1(db)
        for a in actions:
            print(a)
        print(f"Fertig: {len(actions)} Aktionen")
        return 0
    finally:
        db.close()


if __name__ == "__main__":
    raise SystemExit(main())
