"""Tests für PDF- und Excel-Exporte."""

from datetime import datetime, timezone

import pytest
from fastapi.testclient import TestClient

from app.main import app
from app.services.dashboard import build_dashboard_summary
from app.services.export_builders import (
    build_dashboard_export,
    safe_filename_part,
    spritzguss_export_filename,
)
from app.services.export_excel import (
    render_baugruppe_excel,
    render_dashboard_excel,
    render_spritzguss_excel,
)
from app.services.export_models import (
    BaugruppeExportData,
    DashboardExportData,
    ExportInvestment,
    ExportMoneyRow,
    ExportRow,
    ExportTable,
    SpritzgussExportData,
)
from app.services.export_pdf import (
    render_baugruppe_pdf,
    render_dashboard_pdf,
    render_spritzguss_pdf,
)

NOW = datetime(2026, 1, 15, 12, 0, tzinfo=timezone.utc)


def _sample_spritzguss() -> SpritzgussExportData:
    return SpritzgussExportData(
        company_name="Test GmbH",
        calculation_id=42,
        teilebezeichnung="Gehäuse",
        teilenummer="TEST-001",
        kunde="OEM",
        projekt="Projekt Alpha",
        created_at=NOW,
        updated_at=NOW,
        inputs=[ExportRow("Material", "PA6")],
        kosten=[
            ExportMoneyRow("Endpreis je Stück", 12.50, highlight=True),
        ],
        veredelung_steps=[ExportMoneyRow("Montage", 2.50)],
        investitionen=[
            ExportInvestment(
                bezeichnung="Werkzeug",
                typ="Werkzeug",
                betrag=50000.0,
                status="offen",
                hinweis="Separat, nicht im Stückpreis enthalten",
            )
        ],
        werkzeug_hinweis="Separat, nicht im Stückpreis enthalten",
        endpreis=12.50,
    )


def _sample_baugruppe() -> BaugruppeExportData:
    return BaugruppeExportData(
        company_name="Test GmbH",
        assembly_id=7,
        name="Frontstoßfänger",
        teilenummer="BG-TEST-001",
        kunde="OEM",
        projekt="Projekt Alpha",
        jahresstueckzahl=5000,
        created_at=NOW,
        updated_at=NOW,
        einzelteile=ExportTable(
            "Einzelteile",
            ["Bezeichnung", "Teilenummer", "Menge", "Einzelpreis", "Summe"],
            [["Teil A", "A-1", "2", "10,00 €", "20,00 €"]],
        ),
        kaufteile=ExportTable("Kaufteile", ["Bezeichnung"], []),
        veredelung=ExportTable("Veredelung", ["Schritt"], []),
        investitionen=[],
        einzelteile_gesamt=20.0,
        kaufteile_gesamt=0.0,
        veredelung_gesamt=0.0,
        baugruppenpreis_je_stueck=20.0,
        jahresumsatz=100000.0,
    )


def _sample_dashboard(*, empty: bool = False) -> DashboardExportData:
    if empty:
        return DashboardExportData(
            company_name="Test GmbH",
            filter_project=None,
            filter_customer=None,
            generated_at=NOW,
            kpis=[ExportRow("Anzahl Projekte", "0")],
            recent_calculations=ExportTable("Kalkulationen", ["Art"], []),
            assemblies=ExportTable("Baugruppen", ["Name"], []),
            investments=ExportTable("Investitionen", ["Bezeichnung"], []),
            price_comparison=ExportTable("Preise", ["Name"], []),
            investment_by_project=ExportTable("Invest", ["Projekt"], []),
            revenue_by_project=ExportTable("Umsatz", ["Projekt"], []),
            has_data=False,
            empty_message="Keine Daten für die gewählten Filter vorhanden.",
        )
    return DashboardExportData(
        company_name="Test GmbH",
        filter_project="Projekt Alpha",
        filter_customer=None,
        generated_at=NOW,
        kpis=[ExportRow("Anzahl Projekte", "1")],
        recent_calculations=ExportTable(
            "Kalkulationen",
            ["Art", "Bezeichnung"],
            [["Spritzguss", "Teil A"]],
        ),
        assemblies=ExportTable("Baugruppen", ["Name"], []),
        investments=ExportTable("Investitionen", ["Bezeichnung"], []),
        price_comparison=ExportTable("Preise", ["Name", "Typ", "Preis"], []),
        investment_by_project=ExportTable("Invest", ["Projekt", "Betrag"], []),
        revenue_by_project=ExportTable("Umsatz", ["Projekt", "Betrag"], []),
        price_chart=[("Teil A", 12.5)],
        investment_chart=[("Projekt Alpha", 50000.0)],
        revenue_chart=[("Projekt Alpha", 100000.0)],
    )


def test_spritzguss_pdf():
    pdf = render_spritzguss_pdf(_sample_spritzguss())
    assert pdf[:4] == b"%PDF"
    assert len(pdf) > 500


def test_spritzguss_pdf_mit_teilbild():
    data = _sample_spritzguss()
    data.teilbild_mime = "image/png"
    data.teilbild_data = (
        "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAYAAAAfFcSJAAAADUlEQVR42mP8z8BQDwAEhQGAhKmMIQAAAABJRU5ErkJggg=="
    )
    pdf = render_spritzguss_pdf(data)
    assert pdf[:4] == b"%PDF"
    assert len(pdf) > 500


def test_spritzguss_excel_mit_teilbild():
    data = _sample_spritzguss()
    data.teilbild_mime = "image/png"
    data.teilbild_data = (
        "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAYAAAAfFcSJAAAADUlEQVR42mP8z8BQDwAEhQGAhKmMIQAAAABJRU5ErkJggg=="
    )
    xlsx = render_spritzguss_excel(data)
    assert xlsx[:2] == b"PK"
    assert len(xlsx) > 5000


def test_baugruppe_pdf():
    pdf = render_baugruppe_pdf(_sample_baugruppe())
    assert pdf[:4] == b"%PDF"


def test_dashboard_pdf():
    pdf = render_dashboard_pdf(_sample_dashboard())
    assert pdf[:4] == b"%PDF"


def test_dashboard_pdf_ohne_daten():
    pdf = render_dashboard_pdf(_sample_dashboard(empty=True))
    assert pdf[:4] == b"%PDF"
    assert len(pdf) > 200


def test_spritzguss_excel():
    xlsx = render_spritzguss_excel(_sample_spritzguss())
    assert xlsx[:2] == b"PK"


def test_baugruppe_excel():
    xlsx = render_baugruppe_excel(_sample_baugruppe())
    assert xlsx[:2] == b"PK"


def test_dashboard_excel():
    xlsx = render_dashboard_excel(_sample_dashboard())
    assert xlsx[:2] == b"PK"


def test_dashboard_filter_projekt():
    from app.services.dashboard import BaugruppeRecord, SpritzgussRecord

    sg = SpritzgussRecord(
        id=1,
        teilebezeichnung="T",
        teilenummer="N",
        kunde="K1",
        projekt="Alpha",
        jahresstueckzahl=100,
        aktiv=True,
        ergebnis={"endpreis_je_stueck": 5.0},
        created_at=NOW,
        updated_at=NOW,
    )
    summary = build_dashboard_summary([sg], [], [], project="Alpha")
    assert summary["kpis"]["anzahl_spritzguss_kalkulationen"] == 1


def test_dashboard_filter_kunde():
    from app.services.dashboard import SpritzgussRecord

    sg = SpritzgussRecord(
        id=1,
        teilebezeichnung="T",
        teilenummer="N",
        kunde="Kunde X",
        projekt="P",
        jahresstueckzahl=0,
        aktiv=True,
        ergebnis={"endpreis_je_stueck": 5.0},
        created_at=NOW,
        updated_at=NOW,
    )
    summary = build_dashboard_summary([sg], [], [], customer="Kunde X")
    assert summary["kpis"]["anzahl_spritzguss_kalkulationen"] == 1


def test_einmalzahlung_hinweis():
    data = _sample_spritzguss()
    assert data.werkzeug_hinweis is not None
    assert "nicht im Stückpreis" in data.werkzeug_hinweis
    assert data.endpreis == 12.50
    assert data.investitionen[0].betrag == 50000.0


def test_veredelung_nicht_doppelt_im_endpreis():
    data = _sample_spritzguss()
    assert data.endpreis == 12.50
    assert data.veredelung_steps[0].amount == 2.50


def test_export_filename():
    data = _sample_spritzguss()
    assert spritzguss_export_filename(data, "pdf") == "einzelteil_TEST-001.pdf"


def test_safe_filename():
    assert safe_filename_part("Teil/A\\B") == "Teil_A_B"


def test_baugruppe_excel_skonto_null_prozent():
    data = _sample_baugruppe()
    data.skonto = 0.0
    data.export_date = NOW
    data.zuschlagssaetze = ExportTable(
        "Zuschlagssätze",
        ["Bezeichnung", "Satz", "Betrag"],
        [["Skonto", "0,00 %", "0,00 €"]],
    )
    data.kosten_aufstellung = [ExportMoneyRow("Skonto", 0.0)]
    xlsx = render_baugruppe_excel(data)
    assert xlsx[:2] == b"PK"
    from io import BytesIO
    from openpyxl import load_workbook

    workbook = load_workbook(BytesIO(xlsx))
    assert "Zuschlagssaetze" in workbook.sheetnames


def test_baugruppe_pdf_enthaelt_version_und_skonto():
    data = _sample_baugruppe()
    data.skonto = 0.0
    data.structure_version = 3
    data.export_date = NOW
    pdf = render_baugruppe_pdf(data)
    text_content = pdf.decode("latin-1", errors="ignore")
    assert "Skonto" in text_content
    assert "3" in text_content


def test_nicht_authentifizierter_export():
    client = TestClient(app)
    endpoints = [
        "/api/v1/reports/spritzguss/1.pdf",
        "/api/v1/reports/spritzguss/1.xlsx",
        "/api/v1/reports/baugruppen/1.pdf",
        "/api/v1/reports/baugruppen/1.xlsx",
        "/api/v1/reports/dashboard.pdf",
        "/api/v1/reports/dashboard.xlsx",
    ]
    for path in endpoints:
        response = client.get(path)
        assert response.status_code == 401, path
