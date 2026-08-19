"""Phase-D-Tests: Dashboard-Filter, Baugruppendetail und Excel/PDF-Export."""

from __future__ import annotations

import json
from datetime import UTC, date, datetime
from io import BytesIO
from types import SimpleNamespace

import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine, text
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from app.api.router import api_router
from app.core.security import UserRole
from app.database import get_db
from app.dependencies import get_current_user
from app.models.baugruppe import Baugruppe
from app.models.investition import Investition
from app.services.dashboard_assembly import build_assembly_overview
from tests.test_assembly_calculation_phase_c import (
    _create_phase_c_schema,
    _seed_project,
    _top,
)

NOW = datetime(2026, 1, 15, 12, 0, tzinfo=UTC)

PHASE_C_ERGEBNIS = {
    "herstellkosten": 10.57,
    "vvgk": 1.06,
    "selbstkosten": 11.63,
    "gewinn": 1.74,
    "nettoverkaufspreis": 13.37,
    "skonto": 0.0,
    "endpreis_je_stueck": 13.37,
    "markup_applied": True,
    "positions": [
        {
            "position_id": 1,
            "position_type": "PART",
            "sequence": 1,
            "label": None,
            "name_snapshot": "Träger",
            "einzelpreis": 10.57,
            "quantity": 1,
            "quantity_factor": 1,
            "zwischensumme": 10.57,
        }
    ],
    "warnings": [],
}


def _create_phase_d_schema(engine) -> None:
    _create_phase_c_schema(engine)
    statements = [
        "ALTER TABLE spritzguss_kalkulationen ADD COLUMN customer_id INTEGER",
        "ALTER TABLE spritzguss_kalkulationen ADD COLUMN program_id INTEGER",
        "ALTER TABLE spritzguss_kalkulationen ADD COLUMN calculation_year INTEGER",
        "ALTER TABLE spritzguss_kalkulationen ADD COLUMN project_volume FLOAT",
        "ALTER TABLE spritzguss_kalkulationen ADD COLUMN material_id INTEGER",
        "ALTER TABLE spritzguss_kalkulationen ADD COLUMN maschine_id INTEGER",
        "ALTER TABLE spritzguss_kalkulationen ADD COLUMN lohnkosten_id INTEGER",
        "ALTER TABLE spritzguss_kalkulationen ADD COLUMN ergebnis_bloecke TEXT",
        "ALTER TABLE spritzguss_kalkulationen ADD COLUMN notizen TEXT NOT NULL DEFAULT ''",
        """
        CREATE TABLE IF NOT EXISTS baugruppe_spritzguss_zuordnungen (
            id INTEGER PRIMARY KEY,
            baugruppe_id INTEGER NOT NULL,
            spritzguss_kalkulation_id INTEGER NOT NULL,
            menge FLOAT NOT NULL DEFAULT 1,
            reihenfolge INTEGER NOT NULL DEFAULT 1,
            snapshot_preis FLOAT NOT NULL DEFAULT 0,
            snapshot_bezeichnung VARCHAR(255) NOT NULL DEFAULT '',
            snapshot_teilenummer VARCHAR(100) NOT NULL DEFAULT '',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        """,
        """
        CREATE TABLE IF NOT EXISTS investitionen (
            id INTEGER PRIMARY KEY,
            name VARCHAR(255) NOT NULL DEFAULT '',
            investment_type VARCHAR(64) NOT NULL DEFAULT 'Werkzeug',
            payment_type VARCHAR(64) NOT NULL DEFAULT 'Einmalzahlung',
            amount FLOAT NOT NULL DEFAULT 0,
            amortization_volume INTEGER,
            cost_per_piece FLOAT,
            project_id VARCHAR(255) NOT NULL DEFAULT '',
            customer VARCHAR(255) NOT NULL DEFAULT '',
            part_name VARCHAR(255) NOT NULL DEFAULT '',
            part_number VARCHAR(255) NOT NULL DEFAULT '',
            calculation_id INTEGER,
            baugruppe_id INTEGER,
            supplier VARCHAR(255) NOT NULL DEFAULT '',
            order_date DATE,
            delivery_date DATE,
            status VARCHAR(64) NOT NULL DEFAULT 'In Planung',
            description TEXT NOT NULL DEFAULT '',
            included_in_unit_price BOOLEAN NOT NULL DEFAULT 0,
            archived BOOLEAN NOT NULL DEFAULT 0,
            linked_project_id INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        """,
    ]
    with engine.begin() as conn:
        for stmt in statements:
            conn.execute(text(stmt))


def _http_app() -> FastAPI:
    application = FastAPI()
    application.include_router(api_router)
    return application


@pytest.fixture()
def db():
    engine = create_engine(
        "sqlite://",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    _create_phase_d_schema(engine)
    Session = sessionmaker(bind=engine, expire_on_commit=False)
    session = Session()
    yield session
    session.close()


@pytest.fixture()
def client(db):
    application = _http_app()

    def override_get_db():
        yield db

    def override_current_user():
        return SimpleNamespace(
            email="kalkulator@example.com",
            role=UserRole.KALKULATOR.value,
            is_active=True,
        )

    application.dependency_overrides[get_db] = override_get_db
    application.dependency_overrides[get_current_user] = override_current_user
    with TestClient(application) as test_client:
        yield test_client
    application.dependency_overrides.clear()


def _set_ergebnis(obj: Baugruppe, payload: dict) -> None:
    obj.ergebnis = payload


def _seed_dashboard_rows(db):
    _seed_project(db)
    top = _top(db)
    top.name = "Frontmodul"
    top.teilenummer = "FM-1"
    top.kunde = "OEM A"
    top.projekt = "Projekt X"
    top.jahresstueckzahl = 1000
    top.status = "freigegeben"
    top.aktiv = True
    top.created_at = NOW
    top.updated_at = NOW
    _set_ergebnis(top, PHASE_C_ERGEBNIS)
    db.execute(
        text(
            """
            INSERT INTO spritzguss_kalkulationen
            (id, teilebezeichnung, teilenummer, kunde, projekt, jahresstueckzahl, aktiv,
             ergebnis, created_at, updated_at)
            VALUES (501, 'Gehäuse', 'GH-001', 'OEM A', 'Projekt X', 500, 1, :ergebnis, :ts, :ts)
            """
        ),
        {
            "ergebnis": json.dumps({"endpreis_je_stueck": 10.0}),
            "ts": NOW.isoformat(),
        },
    )
    db.add(
        Investition(
            name="Werkzeug A",
            description="Werkzeug-Einmalzahlung",
            part_name="Werkzeug",
            amount=40000.0,
            investment_type="Werkzeug",
            payment_type="Einmalzahlung",
            status="offen",
            project_id="Projekt X",
            customer="OEM A",
            baugruppe_id=top.id,
            supplier="Formenbauer GmbH",
            order_date=date(2026, 1, 10),
            delivery_date=date(2026, 3, 1),
            amortization_volume=10000,
            cost_per_piece=4.0,
            created_at=NOW,
            updated_at=NOW,
        )
    )
    db.commit()
    db.refresh(top)
    return top


def test_http_dashboard_ohne_filter(client, db):
    _seed_dashboard_rows(db)
    response = client.get("/api/v1/dashboard/summary")
    assert response.status_code == 200
    body = response.json()
    assert body["has_data"] is True
    assert body["kpis"]["anzahl_projekte"] == 1
    assert body["kpis"]["anzahl_spritzguss_kalkulationen"] == 1
    assert body["kpis"]["anzahl_baugruppen"] == 1
    assert body["kpis"]["investitionen_gesamt"] == 40000.0
    assert body["kpis"]["umsatzpotenzial_jahr"] == pytest.approx(13370.0 + 5000.0)
    assert body["kpis"]["durchschnitt_preis_pro_stueck"] == pytest.approx(11.69, abs=0.02)
    assert body["recent_calculations"]
    assert body["recent_investments"]
    assert body["cost_structure"]


def test_http_dashboard_mit_kunde_projekt_filter(client, db):
    _seed_dashboard_rows(db)
    response = client.get(
        "/api/v1/dashboard/summary",
        params={"customer": "OEM A", "project": "Projekt X"},
    )
    assert response.status_code == 200
    body = response.json()
    assert body["kpis"]["anzahl_baugruppen"] == 1
    assert body["kpis"]["investitionen_gesamt"] == 40000.0

    empty = client.get(
        "/api/v1/dashboard/summary",
        params={"customer": "Unbekannt", "project": "Nirgends"},
    )
    assert empty.status_code == 200
    empty_body = empty.json()
    assert empty_body["has_data"] is False
    assert empty_body["empty_message"]


def test_http_dashboard_leere_datenmenge(client):
    response = client.get("/api/v1/dashboard/summary")
    assert response.status_code == 200
    body = response.json()
    assert body["has_data"] is False
    assert body["kpis"]["anzahl_projekte"] == 0
    assert "Filter" in body["empty_message"]


def test_baugruppendetail_daten(client, db):
    top = _seed_dashboard_rows(db)
    overview = build_assembly_overview(db, top.id)
    assert overview["einzelteilkosten"] == pytest.approx(10.57)
    assert overview["skonto"] == 0.0
    assert overview["bruttoverkaufspreis"] == pytest.approx(13.37)
    assert overview["jahresumsatz"] == pytest.approx(13370.0)
    assert overview["bom"][0]["bezeichnung"] == "Träger"
    assert any(row["typ"] == "skonto" and row["betrag"] == 0.0 for row in overview["zuschlagssaetze"])

    response = client.get(f"/api/v1/dashboard/assemblies/{top.id}")
    assert response.status_code == 200
    body = response.json()
    assert body["vvgk"] == pytest.approx(1.06)
    assert body["gewinn"] == pytest.approx(1.74)
    assert body["skonto"] == 0.0
    assert body["nettoverkaufspreis"] == pytest.approx(13.37)
    assert body["gesamtsumme"] == pytest.approx(13370.0)


def test_investitionssummen_http(client, db):
    top = _seed_dashboard_rows(db)
    db.add(
        Investition(
            name="Vorrichtung",
            amount=10000.0,
            investment_type="Vorrichtung",
            payment_type="Einmalzahlung",
            status="offen",
            project_id="Projekt X",
            customer="OEM A",
            baugruppe_id=top.id,
            created_at=NOW,
            updated_at=NOW,
        )
    )
    db.commit()
    response = client.get("/api/v1/dashboard/summary")
    body = response.json()
    assert body["kpis"]["investitionen_gesamt"] == 50000.0
    assert body["investment_by_project"][0]["betrag"] == 50000.0


def test_excel_export_http_200_xlsx(client, db):
    top = _seed_dashboard_rows(db)
    response = client.get(f"/api/v1/reports/baugruppen/{top.id}.xlsx")
    assert response.status_code == 200
    assert response.content[:2] == b"PK"
    workbook = load_workbook(BytesIO(response.content))
    assert "Übersicht" in workbook.sheetnames
    assert "BOM" in workbook.sheetnames
    assert "Zuschlagssaetze" in workbook.sheetnames
    sheet = workbook["Zuschlagssaetze"]
    values = [cell.value for row in sheet.iter_rows(max_col=3) for cell in row]
    assert any(value and "Skonto" in str(value) for value in values)


def test_pdf_export_http_200_pdf(client, db):
    top = _seed_dashboard_rows(db)
    response = client.get(f"/api/v1/reports/baugruppen/{top.id}.pdf")
    assert response.status_code == 200
    assert response.content[:4] == b"%PDF"
    text_content = response.content.decode("latin-1", errors="ignore")
    assert "Skonto" in text_content
    assert "Frontmodul" in text_content


def test_export_skonto_null_prozent(client, db):
    top = _seed_dashboard_rows(db)
    xlsx = client.get(f"/api/v1/reports/baugruppen/{top.id}.xlsx")
    workbook = load_workbook(BytesIO(xlsx.content))
    zusammenfassung = workbook["Zusammenfassung"]
    labels = [row[0].value for row in zusammenfassung.iter_rows(min_col=1, max_col=1)]
    assert "Skonto" in labels
    skonto_row = next(row for row in zusammenfassung.iter_rows(min_col=1, max_col=2) if row[0].value == "Skonto")
    assert skonto_row[1].value == 0 or skonto_row[1].value == 0.0


def test_export_aendert_datenbank_nicht(client, db):
    top = _seed_dashboard_rows(db)
    before = db.execute(
        text("SELECT updated_at, ergebnis, jahresstueckzahl FROM baugruppen WHERE id = :id"),
        {"id": top.id},
    ).one()
    assert client.get(f"/api/v1/reports/baugruppen/{top.id}.xlsx").status_code == 200
    assert client.get(f"/api/v1/reports/baugruppen/{top.id}.pdf").status_code == 200
    db.expire_all()
    after = db.execute(
        text("SELECT updated_at, ergebnis, jahresstueckzahl FROM baugruppen WHERE id = :id"),
        {"id": top.id},
    ).one()
    assert after == before
