"""Tests: Baugruppen-Detailkalkulation Excel/PDF-Export inkl. TSV-Überleitung."""

from __future__ import annotations

import json
from datetime import UTC, datetime
from decimal import ROUND_HALF_UP, Decimal
from io import BytesIO

import pytest
from openpyxl import load_workbook
from sqlalchemy import create_engine, text
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from app.models.assembly_position import AssemblyPosition
from app.models.baugruppe import Baugruppe
from app.models.kaufteil import Kaufteil
from app.models.spritzguss_kalkulation import SpritzgussKalkulation
from app.models.spritzguss_veredelung_zuordnung import SpritzgussVeredelungZuordnung
from app.models.zuschlagssatz import Zuschlagssatz
from app.services.assembly_calculation import MarkupRates, PositionCalcInput, calculate_assembly
from app.services.baugruppe_export_detail import build_baugruppe_detail_kalkulation
from app.services.export_builders import build_baugruppe_export
from app.services.export_excel import render_baugruppe_excel
from app.services.export_pdf import render_baugruppe_pdf
from tests.test_assembly_calculation_phase_c import _create_phase_c_schema

NOW = datetime(2026, 3, 20, 12, 0, tzinfo=UTC)


def _money(value: float) -> float:
    return float(Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


@pytest.fixture()
def db():
    engine = create_engine(
        "sqlite://",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    _create_phase_c_schema(engine)
    with engine.begin() as conn:
        for stmt in [
            "ALTER TABLE spritzguss_kalkulationen ADD COLUMN customer_id INTEGER",
            "ALTER TABLE spritzguss_kalkulationen ADD COLUMN program_id INTEGER",
            "ALTER TABLE spritzguss_kalkulationen ADD COLUMN calculation_year INTEGER",
            "ALTER TABLE spritzguss_kalkulationen ADD COLUMN project_volume FLOAT",
            "ALTER TABLE spritzguss_kalkulationen ADD COLUMN material_id INTEGER",
            "ALTER TABLE spritzguss_kalkulationen ADD COLUMN maschine_id INTEGER",
            "ALTER TABLE spritzguss_kalkulationen ADD COLUMN lohnkosten_id INTEGER",
            "ALTER TABLE spritzguss_kalkulationen ADD COLUMN ergebnis_bloecke TEXT",
            "ALTER TABLE spritzguss_kalkulationen ADD COLUMN notizen TEXT NOT NULL DEFAULT ''",
            "ALTER TABLE kaufteile ADD COLUMN customer_id INTEGER",
            "ALTER TABLE kaufteile ADD COLUMN program_id INTEGER",
            "ALTER TABLE kaufteile ADD COLUMN project_id INTEGER",
            """
            CREATE TABLE IF NOT EXISTS customers (
                id INTEGER PRIMARY KEY,
                name VARCHAR(255) NOT NULL DEFAULT '',
                active BOOLEAN DEFAULT 1
            )
            """,
            """
            CREATE TABLE IF NOT EXISTS programs (
                id INTEGER PRIMARY KEY,
                customer_id INTEGER,
                name VARCHAR(255) NOT NULL DEFAULT '',
                active BOOLEAN DEFAULT 1
            )
            """,
            "ALTER TABLE projects ADD COLUMN program_id INTEGER",
            "ALTER TABLE projects ADD COLUMN name VARCHAR(255) DEFAULT ''",
            """
            CREATE TABLE IF NOT EXISTS investitionen (
                id INTEGER PRIMARY KEY,
                name VARCHAR(255) NOT NULL DEFAULT '',
                investment_type VARCHAR(64) NOT NULL DEFAULT 'Werkzeug',
                payment_type VARCHAR(64) NOT NULL DEFAULT 'Einmalzahlung',
                amount FLOAT NOT NULL DEFAULT 0,
                amortization_volume INTEGER,
                cost_per_piece FLOAT,
                project_id VARCHAR(255) NOT NULL DEFAULT '',
                customer VARCHAR(255) NOT NULL DEFAULT '',
                part_name VARCHAR(255) NOT NULL DEFAULT '',
                part_number VARCHAR(255) NOT NULL DEFAULT '',
                calculation_id INTEGER,
                baugruppe_id INTEGER,
                supplier VARCHAR(255) NOT NULL DEFAULT '',
                order_date DATE,
                delivery_date DATE,
                status VARCHAR(64) NOT NULL DEFAULT 'In Planung',
                description TEXT NOT NULL DEFAULT '',
                included_in_unit_price BOOLEAN NOT NULL DEFAULT 0,
                archived BOOLEAN NOT NULL DEFAULT 0,
                linked_project_id INTEGER,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            """,
        ]:
            try:
                conn.execute(text(stmt))
            except Exception:
                pass
    Session = sessionmaker(bind=engine, expire_on_commit=False)
    session = Session()
    yield session
    session.close()


def _seed_rates(db):
    db.add_all(
        [
            Zuschlagssatz(bezeichnung="MGK selbst", satz_prozent=3.0, typ="mgk_kaufteil_selbst", aktiv=True),
            Zuschlagssatz(bezeichnung="MGK OEM", satz_prozent=5.0, typ="mgk_kaufteil_oem", aktiv=True),
            Zuschlagssatz(bezeichnung="FGK", satz_prozent=22.0, typ="fgk", aktiv=True),
            Zuschlagssatz(bezeichnung="VVGK", satz_prozent=10.0, typ="vvgk", aktiv=True),
            Zuschlagssatz(bezeichnung="Gewinn", satz_prozent=15.0, typ="gewinn", aktiv=True),
            Zuschlagssatz(bezeichnung="Skonto", satz_prozent=0.0, typ="skonto", aktiv=True),
        ]
    )
    db.commit()


def _seed_tsv_assembly(db) -> Baugruppe:
    """TSV: Grundträger, Armlehne+Kaschieren, Schalter, ASSY."""
    _seed_rates(db)
    db.execute(text("INSERT INTO customers (id, name) VALUES (1, 'OEM TSV')"))
    db.execute(text("INSERT INTO programs (id, customer_id, name) VALUES (1, 1, 'Programm TSV')"))
    db.execute(
        text("INSERT INTO projects (id, program_id, name) VALUES (100, 1, 'Projekt TSV')")
    )

    db.execute(
        text(
            """
            INSERT INTO veredelungsschritte
            (id, bezeichnung, veredelungsart, taktzeit_s, anzahl_mitarbeiter,
             lohnstundensatz, maschinenstundensatz, verbrauchskosten_je_stueck,
             ausschussquote_pct, fgk_pct, reihenfolge, aktiv)
            VALUES
            (10, 'Kaschieren TPO', 'Kaschieren', 36, 1, 50, 100, 0.5, 1.5, 0, 1, 1),
            (20, 'ASSY Endverbau', 'Montage', 45, 2, 40, 80, 0.2, 1.5, 0, 1, 1)
            """
        )
    )

    db.add_all(
        [
            SpritzgussKalkulation(
                id=1,
                teilebezeichnung="Grundträger",
                teilenummer="GT-1",
                kunde="OEM TSV",
                projekt="Projekt TSV",
                teilegewicht_netto_g=100.0,
                materialpreis_pro_kg=10.0,
                ausschussquote_pct=1.5,
                material_nominierung="selbstnominiert",
                zykluszeit_s=36.0,
                maschinenstundensatz=100.0,
                kavitaeten=1,
                lohnstundensatz=50.0,
                werkzeugkosten_eur=0,
                werkzeug_abrechnungsart="einmalzahlung",
                mgk_pct=3.0,
                fgk_pct=22.0,
                vvgk_pct=10.0,
                gewinn_pct=15.0,
                skonto_pct=0.0,
            ),
            SpritzgussKalkulation(
                id=2,
                teilebezeichnung="Armlehne",
                teilenummer="AL-1",
                kunde="OEM TSV",
                projekt="Projekt TSV",
                teilegewicht_netto_g=200.0,
                materialpreis_pro_kg=8.0,
                ausschussquote_pct=2.5,
                material_nominierung="selbstnominiert",
                zykluszeit_s=40.0,
                maschinenstundensatz=80.0,
                kavitaeten=1,
                lohnstundensatz=40.0,
                werkzeugkosten_eur=0,
                werkzeug_abrechnungsart="einmalzahlung",
                mgk_pct=3.0,
                fgk_pct=22.0,
                vvgk_pct=10.0,
                gewinn_pct=15.0,
                skonto_pct=0.0,
            ),
        ]
    )
    db.flush()
    db.add(
        SpritzgussVeredelungZuordnung(
            kalkulation_id=2,
            veredelungsschritt_id=10,
            reihenfolge=1,
            aktiv=True,
            mengenfaktor=1.0,
        )
    )
    db.add(
        Kaufteil(
            id=1,
            artikelnummer="SW-1",
            bezeichnung="Schalter",
            lieferant="Lieferant A",
            preis=2.50,
            nominierung="selbstnominiert",
            aktiv=True,
        )
    )

    bg = Baugruppe(
        name="TSV",
        teilenummer="TSV-1",
        kunde="OEM TSV",
        projekt="Projekt TSV",
        jahresstueckzahl=41875,
        project_id=100,
        status="aktiv",
        legacy_mode=False,
        structure_version=1,
        created_at=NOW,
        updated_at=NOW,
    )
    db.add(bg)
    db.flush()

    rates = MarkupRates(fgk_pct=22.0, vvgk_pct=10.0, gewinn_pct=15.0, skonto_pct=0.0)
    gt_hk, arm_hk = 4.20, 6.80
    schalter_inkl = 2.58
    result = calculate_assembly(
        assembly_type="TOP_LEVEL",
        positions=[
            PositionCalcInput(
                position_id=1,
                position_type="PART",
                sequence=1,
                quantity=1,
                quantity_factor=1,
                price_basis="COST",
                active=True,
                label="Grundträger",
                name_snapshot="Grundträger",
                cost_snapshot=gt_hk,
                price_snapshot=None,
            ),
            PositionCalcInput(
                position_id=2,
                position_type="PART",
                sequence=2,
                quantity=1,
                quantity_factor=1,
                price_basis="COST",
                active=True,
                label="Armlehne",
                name_snapshot="Armlehne",
                cost_snapshot=arm_hk,
                price_snapshot=None,
            ),
            PositionCalcInput(
                position_id=3,
                position_type="PURCHASED_PART",
                sequence=3,
                quantity=1,
                quantity_factor=1,
                price_basis=None,
                active=True,
                label="Schalter",
                name_snapshot="Schalter",
                cost_snapshot=2.50,
                price_snapshot=schalter_inkl,
            ),
            PositionCalcInput(
                position_id=4,
                position_type="PROCESS",
                sequence=4,
                quantity=1,
                quantity_factor=1,
                price_basis=None,
                active=True,
                label="ASSY Endverbau",
                name_snapshot="ASSY Endverbau",
                cost_snapshot=1.015,
                price_snapshot=None,
                cost_before_scrap=1.00,
                ausschussquote_pct=1.5,
            ),
        ],
        markup_rates=rates,
    )

    db.add_all(
        [
            AssemblyPosition(
                id=1,
                parent_assembly_id=bg.id,
                position_type="PART",
                sequence=1,
                quantity=1.0,
                quantity_factor=1.0,
                part_calculation_id=1,
                price_basis="COST",
                cost_snapshot=gt_hk,
                name_snapshot="Grundträger",
                part_number_snapshot="GT-1",
                active=True,
            ),
            AssemblyPosition(
                id=2,
                parent_assembly_id=bg.id,
                position_type="PART",
                sequence=2,
                quantity=1.0,
                quantity_factor=1.0,
                part_calculation_id=2,
                price_basis="COST",
                cost_snapshot=arm_hk,
                name_snapshot="Armlehne",
                part_number_snapshot="AL-1",
                active=True,
            ),
            AssemblyPosition(
                id=3,
                parent_assembly_id=bg.id,
                position_type="PURCHASED_PART",
                sequence=3,
                quantity=1.0,
                quantity_factor=1.0,
                purchased_part_id=1,
                price_basis=None,
                cost_snapshot=2.50,
                price_snapshot=schalter_inkl,
                name_snapshot="Schalter",
                supplier_snapshot="Lieferant A",
                active=True,
            ),
            AssemblyPosition(
                id=4,
                parent_assembly_id=bg.id,
                position_type="PROCESS",
                sequence=4,
                quantity=1.0,
                quantity_factor=1.0,
                finishing_step_id=20,
                price_basis=None,
                cost_snapshot=1.015,
                name_snapshot="ASSY Endverbau",
                active=True,
            ),
        ]
    )

    ergebnis = {
        "herstellkosten": result.herstellkosten,
        "fgk_basis": result.fgk_basis,
        "fertigungsgemeinkosten": result.fertigungsgemeinkosten,
        "vvgk": result.vvgk,
        "selbstkosten": result.selbstkosten,
        "gewinn": result.gewinn,
        "nettoverkaufspreis": result.nettoverkaufspreis,
        "skonto": result.skonto,
        "endpreis_je_stueck": result.endpreis_je_stueck,
        "markup_applied": True,
        "applied_fgk_pct": 22.0,
        "applied_vvgk_pct": 10.0,
        "applied_gewinn_pct": 15.0,
        "applied_skonto_pct": 0.0,
        "positions": [
            {
                "position_id": p.position_id,
                "position_type": p.position_type,
                "sequence": p.sequence,
                "label": p.label,
                "name_snapshot": p.name_snapshot,
                "einzelpreis": p.einzelpreis,
                "quantity": p.quantity,
                "quantity_factor": p.quantity_factor,
                "zwischensumme": p.zwischensumme,
            }
            for p in result.position_lines
        ],
        "process_yield_details": [
            {
                "position_id": d.position_id,
                "label": d.label,
                "name_snapshot": d.name_snapshot,
                "ausschussquote_pct": d.ausschussquote_pct,
                "vorprodukt_eingang": d.vorprodukt_eingang,
                "process_kosten_vor_ausschuss": d.process_kosten_vor_ausschuss,
                "ausschuss_zuschlag": d.ausschuss_zuschlag,
                "kosten_nach_ausbeute": d.kosten_nach_ausbeute,
            }
            for d in (result.process_yield_details or [])
        ],
        "warnings": [],
    }
    bg.ergebnis = json.dumps(ergebnis)
    db.commit()
    db.refresh(bg)
    return bg


def test_tsv_detail_export_excel_pdf_parity(db):
    bg = _seed_tsv_assembly(db)
    live = json.loads(bg.ergebnis)
    export = build_baugruppe_export(db, bg.id)
    assert export.baugruppenpreis_je_stueck == pytest.approx(live["endpreis_je_stueck"], abs=0.01)
    assert export.jahresstueckzahl == 41875
    assert export.jahresumsatz == pytest.approx(_money(live["endpreis_je_stueck"] * 41875), abs=0.05)
    assert export.program == "Programm TSV"
    assert export.detail is not None

    detail = build_baugruppe_detail_kalkulation(db, bg.id)
    assert len(detail.parts) == 2
    assert len(detail.purchased) == 1
    assert len(detail.processes) == 1

    gt = next(p for p in detail.parts if "Grundträger" in p.bezeichnung)
    assert gt.mgk_pct == 3.0
    assert gt.material_nominierung == "selbstnominiert"
    assert gt.price_basis == "COST"
    assert gt.fgk_pct == 22.0

    arm = next(p for p in detail.parts if "Armlehne" in p.bezeichnung)
    assert arm.process_steps
    kasch = arm.process_steps[0]
    assert kasch.ausschussquote_pct == 1.5
    assert kasch.ausschuss_zuschlag > 0
    assert kasch.vorprodukt_eingang > 0

    sw = detail.purchased[0]
    assert sw.mgk_pct == 3.0
    assert sw.preis_inkl_mgk == 2.58
    assert sw.einkaufspreis == 2.50

    assy = detail.processes[0]
    assert assy.ausschussquote_pct == 1.5
    assert assy.vorprodukt_eingang > assy.kosten_vor_ausschuss
    assert any("Vorprodukt" in h or "Ausbeute" in h for h in assy.hinweise)

    end_line = next(l for l in detail.ueberleitung if l.label == "Endpreis je Stück")
    assert end_line.amount == pytest.approx(live["endpreis_je_stueck"], abs=0.01)
    jahres = next(l for l in detail.ueberleitung if l.label == "Jahresumsatz")
    assert jahres.amount == pytest.approx(export.jahresumsatz, abs=0.05)

    xlsx = render_baugruppe_excel(export)
    assert xlsx[:2] == b"PK"
    wb = load_workbook(BytesIO(xlsx))
    for name in (
        "Deckblatt",
        "Annahmen",
        "Einzelteile",
        "PART_Prozesskette",
        "Kaufteile",
        "ASSY_Prozesskette",
        "Ueberleitung",
        "Zusammenfassung",
    ):
        assert name in wb.sheetnames

    deck = wb["Deckblatt"]
    labels = {row[0].value: row[1].value for row in deck.iter_rows(min_row=4, max_col=2)}
    assert labels["Jahresstückzahl"] == 41875
    assert labels["Programm"] == "Programm TSV"
    assert abs(float(labels["Endpreis je Stück"]) - live["endpreis_je_stueck"]) < 0.05

    annahmen = wb["Annahmen"]
    annahmen_text = " ".join(
        str(c.value) for row in annahmen.iter_rows() for c in row if c.value is not None
    )
    assert "MGK selbstnominiert" in annahmen_text
    assert "3" in annahmen_text

    pdf = render_baugruppe_pdf(export)
    assert pdf[:4] == b"%PDF"
    pdf_text = pdf.decode("latin-1", errors="ignore")
    assert "Endpreis" in pdf_text or "19" in pdf_text
    assert "ASSY" in pdf_text or "Ausbeute" in pdf_text
    assert "Herstellkosten" in pdf_text
    assert "SG&A" in pdf_text or "VVGK" in pdf_text


def test_missing_nominierung_and_rates_visible(db):
    _seed_rates(db)
    db.execute(text("INSERT INTO projects (id) VALUES (100)"))
    kt = Kaufteil(
        id=9,
        artikelnummer="X-1",
        bezeichnung="Ohne Nominierung",
        preis=10.0,
        nominierung=None,
        aktiv=True,
    )
    db.add(kt)
    bg = Baugruppe(
        name="Test",
        teilenummer="T-1",
        jahresstueckzahl=100,
        project_id=100,
        legacy_mode=False,
        ergebnis=json.dumps(
            {
                "herstellkosten": 10.5,
                "vvgk": 1.05,
                "selbstkosten": 11.55,
                "gewinn": 1.73,
                "nettoverkaufspreis": 13.28,
                "skonto": 0.0,
                "endpreis_je_stueck": 13.28,
                "positions": [
                    {
                        "position_id": 1,
                        "position_type": "PURCHASED_PART",
                        "sequence": 1,
                        "name_snapshot": "Ohne Nominierung",
                        "einzelpreis": 10.5,
                        "quantity": 1,
                        "quantity_factor": 1,
                        "zwischensumme": 10.5,
                    }
                ],
                "process_yield_details": [],
                "warnings": [],
            }
        ),
        created_at=NOW,
        updated_at=NOW,
    )
    db.add(bg)
    db.flush()
    db.add(
        AssemblyPosition(
            parent_assembly_id=bg.id,
            position_type="PURCHASED_PART",
            sequence=1,
            quantity=1.0,
            purchased_part_id=9,
            cost_snapshot=10.0,
            price_snapshot=10.0,
            name_snapshot="Ohne Nominierung",
            active=True,
        )
    )
    db.commit()

    detail = build_baugruppe_detail_kalkulation(db, bg.id)
    assert detail.purchased
    assert detail.purchased[0].nominierung is None
    assert detail.purchased[0].hinweise
    assert any("Nominierung" in h or "nominier" in h.lower() for h in detail.purchased[0].hinweise)

    # ohne Raten: löschen
    db.execute(text("DELETE FROM zuschlagssaetze"))
    db.commit()
    detail2 = build_baugruppe_detail_kalkulation(db, bg.id)
    assert any("Zuschlag" in w or "fehlen" in w.lower() for w in detail2.warnings + detail2.assumptions[0].hinweis.split())


def test_sample_baugruppe_renderer_still_works():
    from app.services.export_models import (
        BaugruppeExportData,
        ExportInvestment,
        ExportMoneyRow,
        ExportTable,
    )

    data = BaugruppeExportData(
        company_name="Test",
        assembly_id=1,
        name="X",
        teilenummer="T",
        kunde="K",
        projekt="P",
        jahresstueckzahl=100,
        created_at=NOW,
        updated_at=NOW,
        einzelteile=ExportTable("E", ["A"], []),
        kaufteile=ExportTable("K", ["A"], []),
        veredelung=ExportTable("V", ["A"], []),
        investitionen=[
            ExportInvestment("W", "Werkzeug", 1.0, "offen", "Separat, nicht im Stückpreis enthalten")
        ],
        einzelteile_gesamt=0,
        kaufteile_gesamt=0,
        veredelung_gesamt=0,
        baugruppenpreis_je_stueck=1.0,
        jahresumsatz=100.0,
        kosten_aufstellung=[ExportMoneyRow("HK", 1.0)],
        herstellkosten=1.0,
        selbstkosten=1.1,
        program="Prog",
    )
    assert render_baugruppe_excel(data)[:2] == b"PK"
    assert render_baugruppe_pdf(data)[:4] == b"%PDF"
